import openpyxl
import os


def getTestData(test_case_name):
    dict = {}
    root_path = os.path.dirname(os.path.dirname(os.path.realpath(__file__)))
    data_input_filename = root_path + os.sep + 'Data' + os.sep + 'DataSource.xlsx'

    book = openpyxl.load_workbook(data_input_filename)
    sheet = book.active
    for i in range(1, sheet.max_row + 1):  # to get rows
        if sheet.cell(row=i, column=1).value == test_case_name:
            for j in range(2, sheet.max_column + 1):  # to get columns
                # Dict["lastname"]="shetty
                dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
    return[dict]